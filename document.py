from openpyxl import load_workbook
from pandas import DataFrame
from itertools import islice
import io
import os
import sys
from pdfminer.high_level import extract_text_to_fp

class DocumentSet:
    def __init__(self):
        pass

    def __iter__(self):
        return self


class Document:
    def __init__(self, text, metadata):
        self.text = text
        self.metadata = metadata


class PDFTranscriptDocumentSet(DocumentSet):
    def __init__(self, transcripts_filepath):
        self.filepath = transcripts_filepath

        # Compose self.transcript_filepaths list
        if not os.path.exists(transcripts_filepath):
            print('{} does not exist'.format(transcripts_filepath))
            sys.exit(1)
        self.transcript_filepaths = []
        if os.path.isdir(transcripts_filepath):
            for filename in os.listdir(transcripts_filepath):
                filepath = os.path.join(transcripts_filepath, filename)
                if os.path.isfile(filepath) and filename.lower().endswith('.pdf'):
                    self.transcript_filepaths.append(filepath)
        else:
            self.transcript_filepaths.append(transcripts_filepath)

        # finally, make it an iterable
        self.transcript_filepaths = iter(self.transcript_filepaths)


    def __next__(self):
        """ Iterator to yield Document, where each Transcript is a Document """

        # get the path to the next file
        pdfFile = self.transcript_filepaths.__next__()
        # extract the text
        text = self._extract_text(pdfFile)
        md = self._show_data(pdfFile)
        doc = Document(text=text, metadata=md)

        return doc


    def _extract_text(self, pdf_filepath):
        """ Internal utility function to extract text from a single PDF file """
        with open(pdf_filepath, "rb") as fp:
            text_fp = io.StringIO()
            extract_text_to_fp(fp, text_fp)
            return text_fp.getvalue()


    def _show_data(self, show_file_path):
        show_file_name = os.path.split(show_file_path)[1]

        show_info = {}
        month = show_file_name[0:2]
        day = show_file_name[3:5]
        year = show_file_name[6:10]

        show_info['show_file_path'] = show_file_path
        show_info['show_date'] = month+'/'+day+'/'+year
        show_info['show_id'] = show_file_name[11:14]
        show_info['show_name'] = show_file_name[15:-4] # leave off .PDF
        return show_info


class SFMExtractDocumentSet(DocumentSet):
    def __init__(self, sfmfilepath):
        """ Initialize with the path to a (single) SFM extract Excel file"""
        # must be an .xlsx file!
        wb = load_workbook(filename=sfmfilepath, read_only=True)
        ws = wb['Sheet1']
        data = ws.values

        # Create a DataFrame using the first row as variable/column names
        cols = next(data)
        data = list(data)
        data = (islice(r, 0, None) for r in data)
        df = DataFrame(data, columns=cols)
        self.df_iterrows = df.iterrows()


    def __next__(self):
        """ Iterator to yield Documents, where each Tweet is a Document """

        index, line = self.df_iterrows.__next__()
        text = line['text']
        md = self._tweet_data(line)
        doc = Document(text=text, metadata=md)
        return doc


    def _tweet_data(self, tweet):
        tweet_info = {}
        tweet_info['id'] = tweet['id']
        tweet_info['tweet_url'] = tweet['tweet_url']
        created_at = tweet['parsed_created_at']
        tweet_info['created_date'] = created_at[5:7]+'/'+created_at[8:10]+'/'+created_at[0:5]
        tweet_info['user_screen_name'] = tweet['user_screen_name']
        tweet_info['tweet_type'] = tweet['tweet_type']
        return tweet_info


class EmailExtractDocumentSet(DocumentSet):
    def __init__(self, emailfilepath):
        """ Initialize with the path to a (single) SFM extract Excel file"""
        # must be an .xlsx file!
        wb = load_workbook(filename=emailfilepath, read_only=True)
        ws = wb['Sheet1']
        data = ws.values

        # Create a DataFrame using the first row as variable/column names
        cols = next(data)
        data = list(data)
        data = (islice(r, 0, None) for r in data)
        df = DataFrame(data, columns=cols)
        self.df_iterrows = df.iterrows()


    def __next__(self):
        """ Iterator to yield Documents, where each Tweet is a Document """

        index, line = self.df_iterrows.__next__()
        text = line['Message']
        md = self._email_data(line)
        doc = Document(text=text, metadata=md)
        return doc


    def _email_data(self, email):
        email_info = {}
        email_info['Date'] = email['Date']
        email_info['From'] = email['From']
        email_info['Subject'] = email['Subject']
        return email_info

